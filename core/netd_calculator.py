import math
import os
from collections import Counter
from dataclasses import dataclass
from typing import List, Optional

import openpyxl
import xlrd

from core.logger import AppLogger

log = AppLogger.get(__name__)

WINDOW_SIZE = 12


@dataclass
class FrameResult:
    filename: str
    tbar: float              # selected window mean temperature (°C)
    spatial_noise_mk: float  # sigma of selected window × 1000 (mK)
    netd_mk: float           # spatial_noise_mk / 1.88 (mK)


@dataclass
class MultiFrameResult:
    frames: List[FrameResult]
    mean_spatial_noise: float  # average of all per-frame spatial noises (mK)
    netd_mk: float             # mean_spatial_noise / 1.88  (mK)


class FrameCalculator:
    """
    Processes a single Excel frame for spatial noise calculation.

    Algorithm:
      1. Load sheet into 2D grid; None for empty / non-numeric cells.
      2. Collect all valid values into a flat list.
      3. Compute full-sheet mean (frequency-weighted) — used only to select window.
      4. Slide a 12×12 window across the grid; skip any window that contains a
         non-numeric cell.
      5. Select the window whose mean is closest to the full-sheet mean.
      6. Compute sigma on the 144 values of that selected window.
      7. spatial_noise_mk = sigma × 1000.

    The reported Tbar is the mean of the selected window, not the full-sheet mean.
    """

    def __init__(self, excel_path: str):
        self._path = excel_path
        self._grid: List[List[Optional[float]]] = []

    # ── Public ────────────────────────────────────────────────────────────────

    def load(self) -> None:
        log.info("Loading frame: %s", self._path)
        ext = os.path.splitext(self._path)[1].lower()

        self._grid = []
        if ext == ".xls":
            wb = xlrd.open_workbook(self._path)
            ws = wb.sheet_by_index(0)
            for row_idx in range(ws.nrows):
                parsed_row: List[Optional[float]] = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                        parsed_row.append(None)
                    else:
                        try:
                            parsed_row.append(float(cell.value))
                        except (TypeError, ValueError):
                            parsed_row.append(None)
                self._grid.append(parsed_row)
        else:
            wb = openpyxl.load_workbook(self._path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                parsed_row: List[Optional[float]] = []
                for cell in row:
                    if cell is None:
                        parsed_row.append(None)
                    else:
                        try:
                            parsed_row.append(float(cell))
                        except (TypeError, ValueError):
                            parsed_row.append(None)
                self._grid.append(parsed_row)

        flat = self._flat_valid()
        if not flat:
            raise ValueError(
                f"No valid numeric values found in: {os.path.basename(self._path)}"
            )

        R = len(self._grid)
        C = max((len(r) for r in self._grid), default=0)
        log.info("Loaded %d valid values, grid %d×%d", len(flat), R, C)

    def calculate(self) -> FrameResult:
        if not self._grid:
            raise RuntimeError("Call load() before calculate().")

        flat = self._flat_valid()
        sheet_mean = self._frequency_mean(flat)
        log.debug("Full-sheet mean: %.4f°C  (N=%d) — used for window selection only",
                  sheet_mean, len(flat))

        window_data = self._best_window(sheet_mean)

        tbar = sum(window_data) / len(window_data)
        sigma = self._sigma(window_data, tbar)
        spatial_noise_mk = round(sigma * 1000, 1)
        frame_netd_mk = round(spatial_noise_mk / 1.88, 1)

        filename = os.path.basename(self._path)
        log.info("Frame %s → Tbar=%.4f°C  σ=%.6f  SN=%.1f mK  NETD=%.1f mK",
                 filename, tbar, sigma, spatial_noise_mk, frame_netd_mk)

        return FrameResult(
            filename=filename,
            tbar=round(tbar, 4),
            spatial_noise_mk=spatial_noise_mk,
            netd_mk=frame_netd_mk,
        )

    # ── Private helpers ───────────────────────────────────────────────────────

    def _flat_valid(self) -> List[float]:
        return [v for row in self._grid for v in row if v is not None]

    @staticmethod
    def _frequency_mean(values: List[float]) -> float:
        freq = Counter(values)
        N = len(values)
        return sum(t * f for t, f in freq.items()) / N

    def _best_window(self, sheet_mean: float) -> List[float]:
        WS = WINDOW_SIZE
        grid = self._grid
        R = len(grid)
        C = max((len(row) for row in grid), default=0)

        if R < WS or C < WS:
            raise ValueError(
                f"Sheet dimensions ({R}×{C}) are too small for a {WS}×{WS} window analysis."
            )

        best_data: Optional[List[float]] = None
        best_diff = float("inf")

        for r in range(R - WS + 1):
            for c in range(C - WS + 1):
                window = self._extract_window(r, c)
                if window is None:
                    continue
                w_mean = sum(window) / len(window)
                diff = abs(w_mean - sheet_mean)
                if diff < best_diff:
                    best_diff = diff
                    best_data = window

        if best_data is None:
            raise ValueError(
                "No valid 12×12 window found — all candidate windows contain "
                "at least one non-numeric cell."
            )

        return best_data

    def _extract_window(self, r: int, c: int) -> Optional[List[float]]:
        """Return the 144 values of the WS×WS window at (r, c), or None if any cell is invalid."""
        WS = WINDOW_SIZE
        window: List[float] = []
        for wr in range(WS):
            row = self._grid[r + wr]
            for wc in range(WS):
                idx = c + wc
                if idx >= len(row) or row[idx] is None:
                    return None
                window.append(row[idx])
        return window

    @staticmethod
    def _sigma(data: List[float], tbar: float) -> float:
        """Frequency-weighted sigma on the provided data list."""
        freq = Counter(data)
        N = len(data)
        variance_sum = sum(f * (t - tbar) ** 2 for t, f in freq.items())
        return math.sqrt(variance_sum / N)


class MultiFrameCalculator:
    """
    Orchestrates calculation across multiple Excel frames and aggregates results.

    After all frames are processed:
      mean_spatial_noise = mean(SN1, SN2, ..., SNM)
      NETD               = mean_spatial_noise / 1.88
    """

    def __init__(self, excel_paths: List[str]):
        self._paths = excel_paths

    def calculate_all(self) -> MultiFrameResult:
        if not self._paths:
            raise ValueError("No Excel files provided.")
        if len(self._paths) > 50:
            raise ValueError("Maximum 50 Excel files are supported per run.")

        log.info("Processing %d frame(s)", len(self._paths))
        frames: List[FrameResult] = []

        for path in self._paths:
            calc = FrameCalculator(path)
            calc.load()
            frames.append(calc.calculate())

        mean_sn = round(sum(f.spatial_noise_mk for f in frames) / len(frames), 1)
        netd = round(sum(f.netd_mk for f in frames) / len(frames), 1)

        log.info("Aggregation done — Mean SN=%.1f mK  NETD=%.1f mK", mean_sn, netd)

        return MultiFrameResult(frames=frames, mean_spatial_noise=mean_sn, netd_mk=netd)
